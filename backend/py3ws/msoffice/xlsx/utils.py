from pprint import pprint

import openpyxl
from openpyxl.utils.cell import get_column_letter


def replace_cell_value(wb, sheet_name, search_str, replace_str):
    # wb = openpyxl.load_workbook('example.xlsx')
    # wb.sheetnames
    sheet = wb[sheet_name]
    pprint(sheet)
    amountOfRows = sheet.max_row
    amountOfColumns = sheet.max_column

    for i in range(amountOfColumns):
        for k in range(amountOfRows):
            cell = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
            if cell.find(search_str) != -1:
                newCell = replace_str
                sheet[get_column_letter(i + 1) + str(k + 1)] = newCell

    # wb.save('example_copy.xlsx')


def fill_data(wb, sheet_name, data):
    sheet = wb[sheet_name]

    for key, val in data.items():
        sheet[key] = val


def to_pdf(sheet_name, input_file, output_file):
    from win32com import client
    import win32api
    # input_file = r'C:\Users\thequickblog\Desktop\Python session 2\tqb_sample.xlsx'
    # give your file name with valid path
    # output_file = r'C:\Users\thequickblog\Desktop\Python session 2\tqb_sample_output.pdf'
    # give valid output file name and path
    app = client.DispatchEx("Excel.Application")
    app.Interactive = False
    app.Visible = False

    wb = app.Workbooks.Open(input_file)
    try:
        wb.Worksheets(sheet_name).Activate()
        wb.ActiveSheet.ExportAsFixedFormat(0, output_file)
    except Exception as e:
        print("Failed to convert in PDF format.Please confirm environment meets all the requirements and try again")
    finally:
        wb.Close()
        app.Exit()
